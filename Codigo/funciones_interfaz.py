from PyQt5 import QtCore
from moviepy.video.io.VideoFileClip import VideoFileClip
from nltk.corpus import stopwords
from nltk.tokenize import word_tokenize
import os
import openpyxl as xl
from openpyxl.styles import Alignment, Font
from patrones import * 
from pydub import AudioSegment
from pydub.utils import make_chunks
import speech_recognition as sr

basedir = os.path.dirname(__file__)


#------------ Segunda Ventana ------------#

def cargarVideo(path):
    url = QtCore.QUrl.fromLocalFile(path)
    return url

def cortar(path, valorInicio, valorFin):
    if valorInicio == '' and valorFin == '':
        return path
    else:
        valorInicio =  transformSec(valorInicio)
        valorFin = transformSec(valorFin)

        with VideoFileClip(path) as video:
            new = video.subclip(valorInicio, valorFin)
            nombre_nuevo = path.replace(".mp4", "_cortado.mp4")
            new.write_videofile(nombre_nuevo, audio_codec='aac')
        
        return nombre_nuevo

def transformSec(valor):
    absa = valor.split('.')
    numero = int(absa[0])*60 + int(absa[1])
    return  numero

def get_path():
    with open(os.path.join(basedir,'../Info/path.txt'), 'r') as f:
        path = f.read()
        f.close()
    return path

def delete_path():
    os.remove(os.path.join(basedir,'../Info/path.txt'))

def write_path(path):
    with open(os.path.join(basedir,'../Info/path.txt'), 'w') as f:
        f.write(path)
        f.close()

def nombre_maniobra(path):
    spliteado = path.split('/')
    last = spliteado[len(spliteado)-1]
    last = last.replace('.mp4','')
    return last

#------------ Funciones Tercera Ventana ------------#

# Aisla el audio
def aislarAudio():
    path = get_path()
    video_extraer_audio = VideoFileClip(path)
    #2021_07_06____11_25____1_Subjetiva.ROT 1.mp4
    name = nombre_maniobra(path)
    new_path = os.path.join(basedir,'../Info/', name)
    if not os.path.isdir(new_path):
        os.mkdir(new_path)
    

    audio_path = new_path + '/audio_aislado.wav'
    video_extraer_audio.audio.write_audiofile(audio_path)
    audioFile = sr.AudioFile(audio_path)
    return audio_path
#Chunkea el audio
def chunkeador(audio_path):
    sound_file = AudioSegment.from_wav(audio_path)
    #silencio = split_on_silence(sound_file, min_silence_len=500, silence_thresh=-40 )
    chunks_path = audio_path.replace('/audio_aislado.wav', '/chunks')
    audio_chunks = make_chunks(sound_file, 20000)
    if not os.path.isdir(chunks_path):
        os.mkdir(chunks_path)

    for i, chunk in enumerate(audio_chunks):
        absa = os.path.join(chunks_path , f"chunk{i}.wav")
        chunk.export(absa, format="wav")
    
    delete_path()
    write_path(chunks_path)
    return (i+1)

#Transcribe el audio a texto mediante libreria Speech recognizer
def transcribir(numero_de_chunks):
    reconocedor = sr.Recognizer()
    texto_devolver = ''
    actual_path = get_path()
    for i in range(numero_de_chunks):
        path_chunk = actual_path + "/chunk" + str(i) + ".wav"
        with sr.AudioFile(path_chunk) as source:
            audio_data = reconocedor.record(source)
            text = reconocedor.recognize_google(audio_data, language = "es-ES",  show_all=True)
            
            if (str(text) != "[]"):
                textaux = reconocedor.recognize_google(audio_data, language = "es-ES")
                print("chunk numero:" , i)
                print(str(textaux))
                print("---------")
                texto_devolver = texto_devolver+ " " + str(textaux)
    
    texto_path = actual_path.replace('/chunks','/texto_transcripto.txt')
    actual_path = actual_path.replace(' ', '\ ')
    #os.remove(actual_path)
    instruccion= 'rm -r ' +actual_path
    os.system(instruccion)
        
    with open(texto_path, 'w') as f:
        f.write(texto_devolver)
        f.close()
    return texto_devolver

def pln(texto):
    texto.lower()
    stop_words = stopwords.words('spanish')
    words = word_tokenize(str(texto))

    new_text = ""
    for w in words:
        if w not in stop_words:
            new_text = new_text + " " + w
    if new_text[0] == " ":
        new_text = new_text[1 : : ]
    return new_text

def dividir_texto(texto):
    texto = pln(texto)
    array_return = []
    texto = texto.replace('desembrago', 'desembrague')
    texto = texto.replace('baja', 'bajo')
    action_verbalizada = ["preparo rotonda","rotonda media","rotonda cerca","en rotonda","coche medio","coche cerca","uno libre","uno coches","uno viene","dos libre","dos cohes","tres libre","tres coches","freno","freno suelto","acelero","acelero mantengo","levanto pie acelerador","gira izquierda","girar derecha","recto","intermitente izquierda","intermitente izquierda off","intermitente derecha","intermitente derecha off","lado izquierda libre","lado izquierda ocupado","lado derecha libre","lado derecha ocupado","atrás libre","atrás ocupado","cambio carril izquierda","cambio carril derecha","salgo rotonda","bajo marcha","bajo marcha","bajo marcha","baje marcha","subo marcha","sube marcha","suba marcha","subo marcha","incidente indefinido","embrago","desembrague","mira frente","retrovisor central","mira izquierda","mira retrovisor izquierdo","mira derecha","mira retrovisor derecho","miro detrás","sonido dentro","sonido fuera"]
    action_verbalizada_2 = ["aproximación","rotonda media","rotonda cerca","entro rotonda","coche medio","coche cerca","izquierda libre","izquierda libre","izquierda ocupado","frente libre","frente ocupado","derecha libre","derecha ocupado","piso freno","suelto freno","acelero","mantengo acelerador","suelto acelerador","giro izquierda","giro derecha","recto","izquierda on","izquierda off","derecha on","derecha off","vista izquierda libre","vista izquierda ocupado","vista derecha libre","vista derecha ocupao","atrás libre","atrás ocupado","cambiar izquierda","cambir derecha","salgo rotonda","bajo primera","bajo segunda","bajo tercera","bajo cuarta","subo primera","subo segunda","subo tercera","subo cuarta","incidente indefinido","piso embrague", "suelto embrague","miro delante","miro retrovisor central","miro izquierda","miro retrovisor izquierdo","miro derecha","miro retrovisor derechi","mira detrás","sonido dentro","sonido fuera"]
    array_codes = ["APROX","RND-MD","RND-NR","RND-IN","CAR-MD","CAR-NR","L-FREE","L-FREE","L-BUSY","F-FREE","F-BUSY","R-FREE","R-BUSY","B-ON","B-OFF","T-ON","T-HOLD","T-OFF","TURN-L","TURN-R","STR","LB-ON","LB-OFF","RB-ON","LB-OFF","LV-FREE","LV-BUSY","RV-FREE","RV-BUSY","BK-FREE","BK-BUSY","CHG-L","CHG-R","RND-EXIT","GD","GD","GD","GD","GU","GU","GU","GU","INCIDENT","G-ON","G-OFF","FV","FV-Mirror","LV","LV-Mirror","RV","RV-Mirror","BV","IN-S","OUT-S"]

    for i in range (len(array_codes)):
        str_aux = action_verbalizada[i] + " <" + array_codes[i] + ">"
        str_aux_2 = action_verbalizada_2[i] + " <" + array_codes[i] + ">"

        if (texto.find(action_verbalizada[i]) != -1):
            texto = texto.replace(action_verbalizada[i], str_aux)
        elif (texto.find(action_verbalizada_2[i] )!= -1):
            texto = texto.replace(action_verbalizada_2[i], str_aux_2)
    division_1 = texto.split('>')

    for div_1 in division_1:
        di = div_1.split('<')
        array_return.append(di)
    
    return array_return

def comprobar_verbalizada(verbalizada):

    action_verbalizada = ["preparo rotonda","rotonda media","rotonda cerca","en rotonda","coche medio","coche cerca","uno libre","uno coches","uno viene","dos libre","dos cohes","tres libre","tres coches","freno","freno suelto","acelero","acelero mantengo","levanto pie acelerador","gira izquierda","girar derecha","recto","intermitente izquierda","intermitente izquierda off","intermitente derecha","intermitente derecha off","lado izquierda libre","lado izquierda ocupado","lado derecha libre","lado derecha ocupado","atrás libre","atrás ocupado","cambio carril izquierda","cambio carril derecha","salgo rotonda","bajo marcha","bajo marcha","bajo marcha","baje marcha","subo marcha","sube marcha","suba marcha","subo marcha","incidente indefinido","embrago","desembrague","mira frente","retrovisor central","mira izquierda","mira retrovisor izquierdo","mira derecha","mira retrovisor derecho","miro detrás","sonido dentro","sonido fuera"]
    action_verbalizada_2 = ["aproximación","rotonda media","rotonda cerca","entro rotonda","coche medio","coche cerca","izquierda libre","izquierda libre","izquierda ocupado","frente libre","frente ocupado","derecha libre","derecha ocupado","piso freno","suelto freno","acelero","mantengo acelerador","suelto acelerador","giro izquierda","giro derecha","recto","izquierda on","izquierda off","derecha on","derecha off","vista izquierda libre","vista izquierda ocupado","vista derecha libre","vista derecha ocupao","atrás libre","atrás ocupado","cambiar izquierda","cambiar derecha","salgo rotonda","bajo primera","bajo segunda","bajo tercera","bajo cuarta","subo primera","subo segunda","subo tercera","subo cuarta","incidente indefinido","piso embrague", "suelto embrague","miro delante","miro retrovisor central","miro izquierda","miro retrovisor izquierdo","miro derecha","miro retrovisor derechi","mira detrás","sonido dentro","sonido fuera"]
    array_codes = ["APROX","RND-MD","RND-NR","RND-IN","CAR-MD","CAR-NR","L-FREE","L-FREE","L-BUSY","F-FREE","F-BUSY","R-FREE","R-BUSY","B-ON","B-OFF","T-ON","T-HOLD","T-OFF","TURN-L","TURN-R","STR","LB-ON","LB-OFF","RB-ON","LB-OFF","LV-FREE","LV-BUSY","RV-FREE","RV-BUSY","BK-FREE","BK-BUSY","CHG-L","CHG-R","RND-EXIT","GD","GD","GD","GD","GU","GU","GU","GU","INCIDENT","G-ON","G-OFF","FV","FV-Mirror","LV","LV-Mirror","RV","RV-Mirror","BV","IN-S","OUT-S"]
    for i in range (0, len(action_verbalizada)):
        if verbalizada == action_verbalizada[i]:
            return array_codes[i]
    for i in range (0, len(action_verbalizada_2)):
        if verbalizada == action_verbalizada_2[i]:
            return array_codes[i]
    else:
        return None

def add_patrones(data):
    for i in range(0, len(data)):
        accion = data[i]
        verb = accion[0]
        cod = accion[2]
        vel = accion[4]
        if cod == 'GU' or cod == 'GD':
            chekeo_1 = check_anterior_cambio(data, i)
            chekeo_2 = check_siguiente_cambio(data, i) 
            counter_vel = i
            if chekeo_1 == False and chekeo_2 == False:
                index = i
                embr = ['embrago','','G-ON','','','','','','','','','','']
                data.insert(index, embr)
                index = i+2
                embr = ['desembrague','','G-OFF','','','','','','','','','','']
                data.insert(index, embr)
                counter_vel = i+1
            elif chekeo_1 == False and chekeo_2 == True:
                index = i
                embr = ['embrago','','G-ON','','','','','','','','','','']
                data.insert(index, embr)
                counter_vel = i+1
            elif chekeo_1 == True and chekeo_2 == False:
                index = i+1
                embr = ['desembrague','','G-OFF','','','','','','','','','','']
                data.insert(index, embr)
                counter_vel = i
            if vel == '':
                vel_av = None
                vel_av = velocidad_patron_marcha(verb, cod)
                if vel_av != None:
                    data[counter_vel] = [verb,accion[1],cod,accion[3],vel_av,accion[5],accion[6],accion[7],accion[8],accion[9],accion[10],accion[11],accion[12]]
    
    return data


def check_anterior_cambio(data, i):
    accion_anterior = data[i-1]
    cod_anterior = accion_anterior[2]
    if cod_anterior == 'G-ON':
        return True
    else:
        return False

def check_siguiente_cambio(data, i):
    accion_anterior = data[i+1]
    cod_siguiente = accion_anterior[2]
    if cod_siguiente == 'G-OFF':
        return True
    else:
        return False

def check_positions(data):
    three_codes = ['APROX','RND-IN','RND-EXIT']
    contador_ok = 0
    index_aprox = None
    index_in = None
    index_out = None
    for i in range (0,len(data)):
        d = data[i]
        if d[2] in three_codes:
            contador_ok = contador_ok +1
        if d[2] == three_codes[0]:
            index_aprox = i
        elif d[2] == three_codes[1]:
            index_in = i
        elif d[2] == three_codes[2]:
            index_out = i
    if contador_ok == 3:
        for i in range (0,len(data)):
            accion = data[i]
            if index_aprox <= i < index_in:
                #aprox
                pos_added = [accion[0],accion[1],accion[2],accion[3],accion[4],accion[5],'APROX',accion[7],accion[8],accion[9],accion[10],accion[11],accion[12]]
            elif index_in <= i < index_out:
                #dentro
                pos_added = [accion[0],accion[1],accion[2],accion[3],accion[4],accion[5],'DENTRO',accion[7],accion[8],accion[9],accion[10],accion[11],accion[12]]
            elif index_out <= i :
                #saliendo
                pos_added = [accion[0],accion[1],accion[2],accion[3],accion[4],accion[5],'SALIENDO',accion[7],accion[8],accion[9],accion[10],accion[11],accion[12]]
            data[i] = pos_added
        
    return data

def write_data_file(data):
    path = get_path().replace('/chunks','/procesado.txt')
    data = str(data).replace('], [', '\n-----------------------------\n')
    with open(path, 'w') as f:
        f.write(data)
        f.close()
    
def write_on_excel(data):
    path =  os.path.join(basedir,'../Info/Rotondas supervisadas v7.xlsx')
    excel_file = xl.load_workbook(path)

    #./Info/2021_07_06____11_25____1_Subjetiva.ROT 1/chunks
    absa = (get_path().split('/'))[-2]
    
    rot = absa.split('.')[1].replace('ROT','Rot')
    fech = absa.split('____')[0].replace('_','-')
    nombre = rot + ' ' + fech
    #Number of sheets in the excel file

    add_hoja(excel_file , data, nombre)

    return True

def add_hoja(excel_file, data, nombre):
    excel_file.create_sheet(nombre)
    
    active_sheet = excel_file[nombre]
    excel_file.active = active_sheet

    fecha1 = nombre.split(' ')[2]
    fecha_spliteada = fecha1.split('-')
    fecha = fecha_spliteada[2]+'/' +fecha_spliteada[1]+'/'+ fecha_spliteada[0]
    active_sheet.append((fecha,''))

    ax = nombre.split(' ')[0]
    aux1= ''
    if ax == 'Rot':
        aux1 = 'ROTONDA '
    ax1 = nombre.split(' ')[1]

    asd = aux1 + ax1
    active_sheet.append(('Acciones verbalizadas','x1', 'x2','x3','x4','x5','x6','x7','x8','x9','x10','x11','x12'))
    active_sheet.append((str(asd),'Percepción', 'Acción','Distancia real al ceda el paso (metros)','Velocidad actual','Límite velocidad GoogleMaps API','Posición','Distancia con coche delante 1 = media 2 = cerca','Zi','Zc','Zd','Coche en Carril izquierdo','Coche en Carril derecho'))
    for act in data:
        active_sheet.append(act)

    title_font2 = Font(size=12, bold=True, color= '000000')
    cell_alignment =  Alignment(horizontal='center', vertical='center')
    max_columns = active_sheet.max_column +1

    palero= 0
     #Fixs the column sizes
    for col in active_sheet.columns:
        if palero == 0:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try: # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.1
            active_sheet.column_dimensions[column].width = adjusted_width
        palero = 8

    for col in range(1,max_columns):
        active_sheet.cell(row = 1, column=col).font = title_font2
        active_sheet.cell(row = 2, column=col).font = title_font2
        active_sheet.cell(row = 3, column=col).font = title_font2
        active_sheet.cell(row = 1, column=col).alignment = cell_alignment
        active_sheet.cell(row = 2, column=col).alignment = cell_alignment
        active_sheet.cell(row = 3, column=col).alignment = cell_alignment

    excel_file.save(filename= os.path.join(basedir,'../Info/Rotondas supervisadas v7.xlsx'))

