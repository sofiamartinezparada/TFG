
import openpyxl as xl
import glob
from data_types import Action
from data_types import Rotonda


def get_all_excel_files():
    path = '/Users/sofiamartinezparada/Library/CloudStorage/OneDrive-Personal/UEM/4º/2º SEMESTRE/TFG/Proyecto/Info'+'/*.xlsx'
    array_names = glob.glob(path)
    return array_names

def get_first_row(active_sheet, max_rows):
    for i in range (1, max_rows):
            inicio = active_sheet.cell(row = i, column = 1).value
            if inicio == 'Acciones verbalizadas':
                return i+2

def read_excel(path):
    rotondas = []
    #load the excel file
    excel_file = xl.load_workbook(path)
    #Get the sheet names
    sheets = excel_file.sheetnames

    #Recorremos las hojas
    for i in range (2 , len(sheets)):
        array_actions = []
        actual_sheet_name = excel_file[sheets[i]]
        excel_file.active =  actual_sheet_name
        active_sheet = excel_file.active

        max_rows = active_sheet.max_row +1
        max_colums = active_sheet.max_column +1
        inicio = get_first_row(active_sheet, max_rows)
        #recorremos las filas desde que empiezan las acciones
        numero_verbalizada = None
        numero_percepcion = None
        numero_accion = None
        numero_distancia_ceda_m = None
        numero_vel_actual = None
        numero_limite_api = None
        numero_pos = None
        numero_distancia_coche = None
        numero_coche_izq = None
        numero_coche_der = None
        numero_tr_zi = None
        numero_tr_zc = None
        numero_tr_zd = None
        for col in range (1,max_colums):
            valor = active_sheet.cell(row = 3, column= col).value
            if valor != None:
                if 'OTONDA' in  valor:
                    numero_verbalizada = col
                elif valor == 'Percepción':
                    numero_percepcion = col
                elif valor == 'Acción':
                    numero_accion = col
                elif valor == 'Distancia real al ceda el paso (metros)':
                    numero_distancia_ceda_m = col
                elif valor == 'Velocidad actual':
                    numero_vel_actual = col
                elif valor == 'Límie velocidad GoogleMaps API':
                    numero_limite_api = col
                elif valor == 'Posición':
                    numero_pos = col
                elif 'Distancia con coche delante' in valor:
                    numero_distancia_coche = col
                elif 'Carril izquierdo' in valor:
                    numero_coche_izq = col
                elif 'Carril derecho' in valor:
                    numero_coche_der = col
                elif valor == 'Zi':
                    numero_tr_zi = col
                elif valor == 'Zc':
                    numero_tr_zc = col
                elif valor == 'Zd':
                    numero_tr_zd = col    

        for j in range (4, max_rows):
            verbalizada = None
            if numero_verbalizada != None:
                verbalizada = active_sheet.cell(row = j, column = numero_verbalizada).value
            if verbalizada != None:
                if numero_percepcion != None:
                    percepcion = active_sheet.cell(row = j, column = numero_percepcion).value
                    if percepcion == '-':
                        percepcion = None
                else:
                    percepcion = None
                if numero_accion != None:
                    accion = active_sheet.cell(row = j, column = numero_accion).value
                    if accion == '-':
                        accion = None
                else:
                    accion = None
                if numero_distancia_ceda_m != None:
                    distancia_ceda_m = active_sheet.cell(row = j, column = numero_distancia_ceda_m).value
                    if distancia_ceda_m == '-':
                        distancia_ceda_m = None
                else:
                    distancia_ceda_m = None
                if numero_vel_actual != None:
                    vel_actual = active_sheet.cell(row = j, column = numero_vel_actual).value
                    if vel_actual == '-':
                        vel_actual = None
                else:
                    vel_actual = None
                if numero_limite_api != None:
                    limite_api = active_sheet.cell(row = j, column = numero_limite_api).value
                    if limite_api == '-':
                        limite_api = None
                else:
                    limite_api = None
                if numero_pos != None:
                    pos = active_sheet.cell(row = j, column = numero_pos).value
                else:
                    pos = None
                if numero_distancia_coche != None:
                    distancia_coche = active_sheet.cell(row = j, column = numero_distancia_coche).value
                    if distancia_coche == '-':
                        distancia_coche = None
                else:
                    distancia_coche = None
                if numero_tr_zi != None:
                    tr_zi = active_sheet.cell(row = j, column = numero_tr_zi).value
                    if tr_zi == '-':
                        tr_zi = None
                else:
                    tr_zi = None
                if numero_tr_zc != None:
                    tr_zc = active_sheet.cell(row = j, column = numero_tr_zc).value
                    if tr_zc == '-':
                        tr_zc = None
                else:
                    tr_zc = None
                if numero_tr_zd != None:
                    tr_zd = active_sheet.cell(row = j, column = numero_tr_zd).value
                    if tr_zd == '-':
                        tr_zd = None
                else:
                    tr_zd = None
                if numero_coche_izq != None:
                    coche_izq = active_sheet.cell(row = j, column = numero_coche_izq).value
                    if coche_izq == '-':
                        coche_izq = None
                else:
                    coche_izq = None
                if numero_coche_der != None:
                    coche_der = active_sheet.cell(row = j, column = numero_coche_der).value
                    if coche_der == '-':
                        coche_der = None
                else:
                    coche_der = None

                action = Action(verbalizada, percepcion, accion, distancia_ceda_m, vel_actual, limite_api, pos, distancia_coche, tr_zi, tr_zc, tr_zd, coche_izq, coche_der)
                array_actions.append(action)

        fecha = active_sheet.cell(row = 1, column = 1).value
        fecha = str(fecha).split(' ')[0]
        rotonda = Rotonda(fecha,array_actions)
        rotondas.append(rotonda)
    return rotondas

def main_read():
    paths = get_all_excel_files()
    for path in paths:
        rotondas = read_excel(path)
    return rotondas
